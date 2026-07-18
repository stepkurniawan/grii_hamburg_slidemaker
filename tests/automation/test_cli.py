from io import BytesIO

from openpyxl import Workbook

from grii_slide_maker.automation import cli


def build_workbook_bytes():
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "info"
    schedule["B2"] = "Sun"
    schedule["B3"] = 9
    schedule["B4"] = 30
    service_order = workbook.create_sheet("dashboard")
    service_order["A185"] = "Songs"
    service_order["B185"] = "161, 320, 93, 169"
    service_order["A187"] = "Bible Reading"
    service_order["B187"] = "Genesis 1:2-3"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_generate_command_builds_and_uploads(monkeypatch):
    calls = []
    monkeypatch.setenv("GOOGLE_SHEET_MASTER_WARTA_ID", "excel-file")
    monkeypatch.setattr(
        cli,
        "download_excel_file_to_memory",
        lambda file_id: build_workbook_bytes(),
    )

    def fake_build_service_slides(service_order, **kwargs):
        calls.append(("songs", [song.value for song in service_order.songs.worship_songs]))
        kwargs["binary_output_file"].write(b"pptx-bytes")

    monkeypatch.setattr(cli, "build_service_slides", fake_build_service_slides)
    monkeypatch.setattr(cli, "sunday_date", lambda formatted: "20260628")
    monkeypatch.setattr(
        cli,
        "upload_or_replace_file",
        lambda **kwargs: calls.append(("upload", kwargs)) or "uploaded-id",
    )

    exit_code = cli.main(["generate"])

    assert exit_code == 0
    assert calls[0] == ("songs", ["161", "320", "93", "169"])
    assert calls[1][0] == "upload"
    assert calls[1][1]["filename"] == "20260628.pptx"
    assert calls[1][1]["content"] == b"pptx-bytes"


def test_generate_command_prints_workbook_validation_error(monkeypatch, capsys):
    workbook_bytes = build_workbook_bytes()
    workbook = cli.load_workbook_from_bytes(workbook_bytes)
    workbook["dashboard"]["B187"] = "Deutronomy 1:1-2"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    monkeypatch.setenv("GOOGLE_SHEET_MASTER_WARTA_ID", "excel-file")
    monkeypatch.setattr(cli, "download_excel_file_to_memory", lambda file_id: output)

    exit_code = cli.main(["generate"])

    assert exit_code == 1
    assert "Unknown English Bible book: Deutronomy" in capsys.readouterr().err


def test_print_cron_command_prints_managed_block(monkeypatch, capsys):
    monkeypatch.setenv("GOOGLE_SHEET_MASTER_WARTA_ID", "excel-file")
    monkeypatch.setattr(
        cli,
        "download_excel_file_to_memory",
        lambda file_id: build_workbook_bytes(),
    )

    exit_code = cli.main(["print-cron", "--command", "generate-command"])

    assert exit_code == 0
    assert "30 9 * * 0 generate-command" in capsys.readouterr().out


def test_default_generate_command_uses_absolute_uv_path(monkeypatch):
    monkeypatch.setattr(cli.os, "getcwd", lambda: "/repo path")
    monkeypatch.setattr(cli.shutil, "which", lambda command: "/home/user/.local/bin/uv")

    command = cli.default_generate_command()

    assert command == (
        "cd '/repo path' && /home/user/.local/bin/uv run grii-slide-auto generate"
    )


def test_sync_cron_command_replaces_existing_block(monkeypatch):
    written = []
    monkeypatch.setenv("GOOGLE_SHEET_MASTER_WARTA_ID", "excel-file")
    monkeypatch.setattr(
        cli,
        "download_excel_file_to_memory",
        lambda file_id: build_workbook_bytes(),
    )
    monkeypatch.setattr(
        cli,
        "read_crontab",
        lambda: "# BEGIN grii-slide-auto\nold\n# END grii-slide-auto\n",
    )
    monkeypatch.setattr(cli, "write_crontab", lambda content: written.append(content))

    exit_code = cli.main(["sync-cron", "--command", "generate-command"])

    assert exit_code == 0
    assert written == [
        "# BEGIN grii-slide-auto\n"
        "30 9 * * 0 generate-command\n"
        "# END grii-slide-auto\n"
    ]


def test_sync_cron_command_removes_block_when_excel_disables_automation(monkeypatch):
    workbook_bytes = build_workbook_bytes()
    workbook = cli.load_workbook_from_bytes(workbook_bytes)
    workbook["info"]["B1"] = "no"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    written = []
    monkeypatch.setenv("GOOGLE_SHEET_MASTER_WARTA_ID", "excel-file")
    monkeypatch.setattr(cli, "download_excel_file_to_memory", lambda file_id: output)
    monkeypatch.setattr(
        cli,
        "read_crontab",
        lambda: (
            "0 1 * * * backup\n"
            "# BEGIN grii-slide-auto\n"
            "old\n"
            "# END grii-slide-auto\n"
        ),
    )
    monkeypatch.setattr(cli, "write_crontab", lambda content: written.append(content))

    exit_code = cli.main(["sync-cron", "--command", "generate-command"])

    assert exit_code == 0
    assert written == ["0 1 * * * backup\n"]
