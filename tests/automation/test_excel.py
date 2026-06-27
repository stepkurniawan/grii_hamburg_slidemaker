from io import BytesIO

from openpyxl import Workbook
import pytest

from grii_slide_maker.automation.excel import (
    load_workbook_from_bytes,
    parse_automation_enabled,
    parse_schedule,
    parse_service_order,
)
from grii_slide_maker.config import Settings


def workbook_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_workbook():
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "info"
    service_order = workbook.create_sheet("dashboard")
    return workbook, schedule, service_order


def test_parse_schedule_from_fixed_cells():
    workbook, schedule, _ = build_workbook()
    schedule["B1"] = "yes"
    schedule["B2"] = "Sun"
    schedule["B3"] = 9
    schedule["B4"] = 30

    loaded = load_workbook_from_bytes(workbook_bytes(workbook))
    parsed_schedule = parse_schedule(loaded, Settings())

    assert parsed_schedule.cron_expression() == "30 9 * * 0"


def test_parse_automation_enabled_from_info_sheet():
    workbook, schedule, _ = build_workbook()
    schedule["B1"] = "no"

    loaded = load_workbook_from_bytes(workbook_bytes(workbook))

    assert parse_automation_enabled(loaded, Settings()) is False


def test_parse_service_order_from_dashboard_labels():
    workbook, _, service_order = build_workbook()
    service_order["A1"] = "Songs"
    service_order["B1"] = "Header next to songs"
    service_order["A174"] = "THIS WEEK"
    service_order["A185"] = "Songs"
    service_order["B185"] = "161, 320, 93, 169"
    service_order["A186"] = "Holy Communion"
    service_order["B186"] = "94"
    service_order["A187"] = "Bible Reading"
    service_order["B187"] = "Genesis 1:2-3, 1 Kings 1:1-2"
    service_order["A188"] = "Preacher"
    service_order["B188"] = "Pdt.Billy Kristanto"

    loaded = load_workbook_from_bytes(workbook_bytes(workbook))
    parsed_order = parse_service_order(loaded, Settings())

    assert [song.value for song in parsed_order.songs.worship_songs] == [
        "161",
        "320",
        "93",
        "169",
    ]
    assert parsed_order.songs.holy_communion_song.value == "94"
    assert parsed_order.pastor.name == "Billy Kristanto"
    assert parsed_order.pastor.title_id == "Pdt."
    assert parsed_order.pastor.title_de_or_en == "Rev."
    assert [ref.as_reference_text() for ref in parsed_order.bible_references] == [
        "Genesis 1:2-3",
        "1 Kings 1:1-2",
    ]


def test_parse_service_order_falls_back_when_this_week_marker_is_missing():
    workbook, _, service_order = build_workbook()
    service_order["A10"] = "Songs"
    service_order["B10"] = "161, 320, 93, 169"
    service_order["A11"] = "Bible Reading"
    service_order["B11"] = "Genesis 1:2-3"

    loaded = load_workbook_from_bytes(workbook_bytes(workbook))
    parsed_order = parse_service_order(loaded, Settings())

    assert [song.value for song in parsed_order.songs.worship_songs] == [
        "161",
        "320",
        "93",
        "169",
    ]


def test_parse_service_order_uses_default_pastor_when_empty():
    workbook, _, service_order = build_workbook()
    service_order["A9"] = "THIS WEEK"
    service_order["A10"] = "Songs"
    service_order["B10"] = "161, 320, 93, 169"
    service_order["A11"] = "Bible Reading"
    service_order["B11"] = "Genesis 1:2-3"

    loaded = load_workbook_from_bytes(workbook_bytes(workbook))
    parsed_order = parse_service_order(loaded, Settings())

    assert parsed_order.pastor.name == "Billy Kristanto"
    assert parsed_order.pastor.title_id == "Pdt."
    assert parsed_order.pastor.title_de_or_en == "Rev."


def test_parse_service_order_rejects_non_text_pastor_name():
    workbook, _, service_order = build_workbook()
    service_order["A9"] = "THIS WEEK"
    service_order["A10"] = "Songs"
    service_order["B10"] = "161, 320, 93, 169"
    service_order["A11"] = "Bible Reading"
    service_order["B11"] = "Genesis 1:2-3"
    service_order["A12"] = "Preacher"
    service_order["B12"] = 123

    loaded = load_workbook_from_bytes(workbook_bytes(workbook))

    with pytest.raises(ValueError, match="Pastor name .* must be text"):
        parse_service_order(loaded, Settings())
