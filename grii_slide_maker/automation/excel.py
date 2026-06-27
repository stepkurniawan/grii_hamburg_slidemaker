"""Read automation schedule and service-order inputs from Excel workbooks."""

from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from grii_slide_maker.automation.schedule import AutomationSchedule
from grii_slide_maker.config import Settings
from grii_slide_maker.models import OrderOfMass

DEFAULT_PASTOR = "Pdt. Billy Kristanto"
DEFAULT_PASTOR_TITLE = "Rev."


def load_workbook_from_bytes(workbook_bytes: BytesIO) -> Workbook:
    workbook_bytes.seek(0)
    return load_workbook(workbook_bytes, data_only=True)


def parse_schedule(workbook: Workbook, settings: Settings) -> AutomationSchedule:
    day = _required_cell(
        workbook,
        sheet_name=settings.AUTOMATION_SCHEDULE_SHEET_NAME,
        cell_reference=settings.AUTOMATION_CRON_DAY_CELL,
        label="cron day",
    )
    hour = _required_int_cell(
        workbook,
        sheet_name=settings.AUTOMATION_SCHEDULE_SHEET_NAME,
        cell_reference=settings.AUTOMATION_CRON_HOUR_CELL,
        label="cron hour",
    )
    minute = _required_int_cell(
        workbook,
        sheet_name=settings.AUTOMATION_SCHEDULE_SHEET_NAME,
        cell_reference=settings.AUTOMATION_CRON_MINUTE_CELL,
        label="cron minute",
    )

    return AutomationSchedule(day=str(day), hour=hour, minute=minute)


def parse_automation_enabled(workbook: Workbook, settings: Settings) -> bool:
    value = _optional_cell(
        workbook,
        sheet_name=settings.AUTOMATION_SCHEDULE_SHEET_NAME,
        cell_reference=settings.AUTOMATION_ENABLED_CELL,
    )
    if value in (None, ""):
        return True

    normalized_value = str(value).strip().lower()
    if normalized_value in {"yes", "y", "true", "1"}:
        return True
    if normalized_value in {"no", "n", "false", "0"}:
        return False

    raise ValueError(
        f"Automation enabled value must be yes or no in "
        f"{settings.AUTOMATION_SCHEDULE_SHEET_NAME}!{settings.AUTOMATION_ENABLED_CELL}"
    )


def parse_service_order(workbook: Workbook, settings: Settings) -> OrderOfMass:
    sheet_name = settings.AUTOMATION_SERVICE_SHEET_NAME
    song_numbers = _required_labeled_value(
        workbook,
        sheet_name,
        label=settings.AUTOMATION_SONG_NUMBERS_LABEL,
        value_label=settings.AUTOMATION_SONG_NUMBERS_LABEL,
        section_label=settings.AUTOMATION_DASHBOARD_SECTION_LABEL,
    )
    bible_verses = _required_labeled_value(
        workbook,
        sheet_name,
        label=settings.AUTOMATION_BIBLE_VERSES_LABEL,
        value_label=settings.AUTOMATION_BIBLE_VERSES_LABEL,
        section_label=settings.AUTOMATION_DASHBOARD_SECTION_LABEL,
    )
    pastor_name = (
        _optional_labeled_value(
            workbook,
            sheet_name,
            label=settings.AUTOMATION_PASTOR_NAME_LABEL,
            section_label=settings.AUTOMATION_DASHBOARD_SECTION_LABEL,
        )
        or DEFAULT_PASTOR
    )
    holy_communion_song = _optional_labeled_value(
        workbook,
        sheet_name,
        label=settings.AUTOMATION_HOLY_COMMUNION_SONG_LABEL,
        section_label=settings.AUTOMATION_DASHBOARD_SECTION_LABEL,
    )
    if not isinstance(pastor_name, str):
        raise ValueError(
            f"Pastor name next to '{settings.AUTOMATION_PASTOR_NAME_LABEL}' "
            f"in sheet '{sheet_name}' must be text"
        )

    return OrderOfMass.model_validate(
        {
            "song_numbers": _stringify(song_numbers),
            "pastor_name": _normalize_pastor_name(pastor_name),
            "bible_verses": _stringify(bible_verses),
            "pastor_title": DEFAULT_PASTOR_TITLE,
            "holy_communion_song_number": _stringify(holy_communion_song)
            if holy_communion_song
            else None,
        }
    )


def _required_cell(
    workbook: Workbook,
    sheet_name: str,
    cell_reference: str,
    label: str,
) -> Any:
    value = _optional_cell(workbook, sheet_name, cell_reference)
    if value in (None, ""):
        raise ValueError(f"Missing {label} in {sheet_name}!{cell_reference}")
    return value


def _optional_cell(workbook: Workbook, sheet_name: str, cell_reference: str) -> Any:
    try:
        worksheet = workbook[sheet_name]
    except KeyError as error:
        raise ValueError(f"Workbook is missing sheet '{sheet_name}'") from error

    value = worksheet[cell_reference].value
    if isinstance(value, str):
        return value.strip()
    return value


def _required_labeled_value(
    workbook: Workbook,
    sheet_name: str,
    label: str,
    value_label: str,
    *,
    section_label: str | None = None,
) -> Any:
    value = _optional_labeled_value(
        workbook,
        sheet_name,
        label,
        section_label=section_label,
    )
    if value in (None, ""):
        raise ValueError(f"Missing {value_label} next to '{label}' in sheet '{sheet_name}'")
    return value


def _optional_labeled_value(
    workbook: Workbook,
    sheet_name: str,
    label: str,
    *,
    section_label: str | None = None,
) -> Any:
    try:
        worksheet = workbook[sheet_name]
    except KeyError as error:
        raise ValueError(f"Workbook is missing sheet '{sheet_name}'") from error

    min_row = _find_section_start_row(worksheet, section_label) or 1
    expected_label = _normalize_label(label)
    for row in worksheet.iter_rows(min_row=min_row):
        for cell in row:
            if _normalize_label(cell.value) != expected_label:
                continue
            value = worksheet.cell(row=cell.row, column=cell.column + 1).value
            if isinstance(value, str):
                return value.strip()
            return value

    return None


def _find_section_start_row(worksheet: Any, section_label: str | None) -> int | None:
    if not section_label:
        return None

    expected_label = _normalize_label(section_label)
    for row in worksheet.iter_rows():
        for cell in row:
            if _normalize_label(cell.value) == expected_label:
                return cell.row

    return None


def _required_int_cell(
    workbook: Workbook,
    sheet_name: str,
    cell_reference: str,
    label: str,
) -> int:
    value = _required_cell(workbook, sheet_name, cell_reference, label)
    try:
        return int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{label.title()} must be an integer") from error


def _stringify(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_label(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _normalize_pastor_name(value: str) -> str:
    # Normalize pastor names so abbreviations like "Pdt." or "Rev." have a single space
    # after the period, e.g. "Pdt.Billy" becomes "Pdt. Billy".
    return re.sub(r"^([A-Za-z]+)\.\s*(\S)", r"\1. \2", value).strip()
